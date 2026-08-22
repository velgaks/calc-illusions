"""Build the public SESH dashboard datasets from the two source workbooks.

The raw workbooks stay in data/raw and are never published.  The browser gets a
small, dictionary-coded subset containing only the variables used by the public
dashboard.  Household keys are replaced by sequential cluster ids.
"""

from __future__ import annotations

import argparse
import json
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
OUT_DIR = ROOT / "public" / "data"

HOUSEHOLD_GLOB = "*SESH Household*.xlsx"
PEOPLE_GLOB = "*SESHS IND*.xlsx"

FIELDWORK = "2023-12/2024-02"
LICENSE = "CC BY-NC-SA 4.0"
SOURCE_URL = "https://www.unicef.org/ukraine/en/documents/socio-economic-status-of-households"


@dataclass(frozen=True)
class Variable:
    source: str
    kind: str
    topic: str
    label_uk: str
    label_en: str
    universe_uk: str
    universe_en: str
    transform: Callable[[dict[str, Any]], Any] | None = None


def number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    return None


def yes_no(value: Any) -> str | None:
    if value == "Так":
        return "yes"
    if value == "Ні":
        return "no"
    if value in {"Не знаю", "Не знаю / не пам’ятаю"}:
        return "unknown"
    return None


def personal_income(row: dict[str, Any]) -> float | None:
    if row.get("IND_V1_1") == "Не маю індивідуального доходу":
        return 0.0
    return number(row.get("IND_V1_1_1"))


def idp_status(row: dict[str, Any]) -> str | None:
    if row.get("IND_A19") == "Так":
        return "yes"
    if row.get("IND_A19") == "Ні":
        return "no"
    if row.get("IND_A15") == "Проживав ще до початку гібридної війни з рф на Сході України (до 20 лютого 2014 року)":
        return "no"
    if row.get("IND_A15") == "Не знаю / не пам’ятаю":
        return "unknown"
    return None


PEOPLE_VARIABLES: dict[str, Variable] = {
    "age": Variable("IND_A2", "numeric", "demography", "Вік", "Age", "Усі члени домогосподарств", "All household members", lambda r: number(r.get("IND_A2"))),
    "sex": Variable("IND_A3", "categorical", "demography", "Стать", "Sex", "Усі члени домогосподарств", "All household members"),
    "region": Variable("OBLAST", "categorical", "demography", "Область проживання", "Region of residence", "Усі члени домогосподарств", "All household members"),
    "settlement": Variable("SETTL_TYPE", "categorical", "demography", "Тип поселення", "Settlement type", "Усі члени домогосподарств", "All household members"),
    "marital": Variable("IND_A6", "categorical", "demography", "Сімейний стан", "Marital status", "Населення віком 15+", "People aged 15+"),
    "education": Variable("IND_A7", "categorical", "demography", "Освіта", "Education", "Населення віком 15+", "People aged 15+"),
    "moved_since_2014": Variable("IND_A15", "categorical", "demography", "Переїзд після 2014 року", "Moved since 2014", "Усі члени домогосподарств", "All household members"),
    "idp": Variable("IND_A19", "categorical", "demography", "Статус ВПО", "IDP status", "Усі члени домогосподарств", "All household members", idp_status),
    "abroad_since_2022": Variable("IND_A21", "categorical", "demography", "Виїзд за кордон після 24.02.2022", "Stayed abroad since 24 Feb 2022", "Усі члени домогосподарств", "All household members", lambda r: yes_no(r.get("IND_A21"))),
    "household_size": Variable("hh_size", "numeric", "demography", "Розмір домогосподарства", "Household size", "Усі члени домогосподарств", "All household members", lambda r: number(r.get("hh_size"))),
    "children_count": Variable("dity_n", "numeric", "demography", "Дітей до 18 років у домогосподарстві", "Children under 18 in the household", "Усі члени домогосподарств", "All household members", lambda r: number(r.get("dity_n"))),
    "work_status": Variable("IND_B1", "categorical", "employment", "Економічний статус", "Economic status", "Населення віком 15+", "People aged 15+"),
    "work_arrangement": Variable("IND_B3_1", "categorical", "employment", "Тип основної роботи", "Main work arrangement", "Ті, хто працювали або тимчасово не працювали", "People working or temporarily absent"),
    "sector": Variable("IND_B5_2", "categorical", "employment", "Сектор економіки", "Economic sector", "Зайняте населення", "Employed people"),
    "occupation_group": Variable("IND_B6_2", "categorical", "employment", "Професійна група", "Occupation group", "Зайняте населення", "Employed people"),
    "weekly_hours": Variable("IND_B9_1", "numeric", "employment", "Годин роботи на тиждень", "Weekly working hours", "Зайняте населення", "Employed people", lambda r: number(r.get("IND_B9_1"))),
    "job_search": Variable("IND_B21", "categorical", "employment", "Пошук роботи", "Looking for work", "Непрацюючі, яким поставили запитання", "Non-working people asked the question"),
    "income_source": Variable("IND_V1_1", "categorical", "income", "Основне джерело особистого доходу", "Main personal income source", "Населення віком 15+", "People aged 15+"),
    "personal_income": Variable("IND_V1_1_1", "numeric", "income", "Основний особистий дохід на місяць", "Main personal monthly income", "Населення віком 15+", "People aged 15+", personal_income),
}


HOUSEHOLD_VARIABLES: dict[str, Variable] = {
    "region": Variable("OBLAST", "categorical", "demography", "Область", "Region", "Усі домогосподарства", "All households"),
    "settlement": Variable("SETTL_TYPE", "categorical", "demography", "Тип поселення", "Settlement type", "Усі домогосподарства", "All households"),
    "household_size": Variable("hh_size", "numeric", "demography", "Розмір домогосподарства", "Household size", "Усі домогосподарства", "All households", lambda r: number(r.get("hh_size"))),
    "children_count": Variable("dity_n", "numeric", "demography", "Кількість дітей до 18 років", "Children under 18", "Усі домогосподарства", "All households", lambda r: number(r.get("dity_n"))),
    "hh_income_total": Variable("V1", "numeric", "income", "Сукупний дохід домогосподарства на місяць", "Total household monthly income", "Усі домогосподарства", "All households", lambda r: number(r.get("V1"))),
    "hh_income_per_capita": Variable("V1", "numeric", "income", "Дохід домогосподарства на одну особу", "Household income per person", "Усі домогосподарства", "All households", lambda r: (number(r.get("V1")) / number(r.get("hh_size"))) if number(r.get("V1")) is not None and number(r.get("hh_size")) not in {None, 0} else None),
    "income_change": Variable("V2", "categorical", "income", "Дохід порівняно з 2021 роком", "Income compared with 2021", "Усі домогосподарства", "All households"),
    "income_adequacy": Variable("V15", "categorical", "income", "Купівельна спроможність доходу", "Income purchasing power", "Усі домогосподарства", "All households"),
    "relative_wealth": Variable("V16", "categorical", "income", "Достаток порівняно із середнім", "Wealth relative to average", "Усі домогосподарства", "All households"),
    "debt": Variable("V5", "categorical", "income", "Борги або кредити", "Debt or loans", "Усі домогосподарства", "All households"),
    "used_savings": Variable("V8", "categorical", "income", "Використовували заощадження", "Used savings", "Усі домогосподарства", "All households"),
    "uses_credit": Variable("V11", "categorical", "income", "Користуються позиками", "Uses credit", "Усі домогосподарства", "All households"),
    "financial_stability": Variable("V13", "categorical", "income", "Фінансова стабільність", "Financial stability", "Усі домогосподарства", "All households"),
    "housing_tenure": Variable("B2", "categorical", "living", "Право користування житлом", "Housing tenure", "Усі домогосподарства", "All households"),
    "dwelling_type": Variable("B3", "categorical", "living", "Тип житла", "Dwelling type", "Усі домогосподарства", "All households"),
    "area_total": Variable("B4", "numeric", "living", "Загальна площа житла", "Total dwelling area", "Домогосподарства, що назвали площу", "Households reporting dwelling area", lambda r: number(r.get("B4"))),
    "area_per_capita": Variable("B4", "numeric", "living", "Площа житла на одну особу", "Dwelling area per person", "Домогосподарства, що назвали площу", "Households reporting dwelling area", lambda r: (number(r.get("B4")) / number(r.get("hh_size"))) if number(r.get("B4")) is not None and number(r.get("hh_size")) not in {None, 0} else None),
    "rooms": Variable("B5", "numeric", "living", "Кількість кімнат", "Number of rooms", "Домогосподарства, що назвали кількість кімнат", "Households reporting rooms", lambda r: number(r.get("B5"))),
    "adequate_area": Variable("B7", "categorical", "living", "Чи достатньо площі житла", "Whether dwelling area is sufficient", "Усі домогосподарства", "All households"),
    "winter_temperature": Variable("B9_1", "categorical", "living", "Комфортна температура взимку", "Comfortable winter temperature", "Усі домогосподарства", "All households"),
    "internet": Variable("B17", "categorical", "living", "Інтернет удома", "Internet at home", "Усі домогосподарства", "All households"),
    "car_count": Variable("B18_4_1", "numeric", "living", "Кількість автомобілів", "Number of cars", "Домогосподарства, яким поставили запитання", "Households asked the question", lambda r: number(r.get("B18_4_1"))),
    "public_transport": Variable("B19", "categorical", "living", "Доступ до громадського транспорту", "Access to public transport", "Усі домогосподарства", "All households"),
    "unexpected_expense": Variable("E1_1", "categorical", "living", "Можуть покрити несподівані витрати", "Can cover an unexpected expense", "Усі домогосподарства", "All households"),
    "annual_holiday": Variable("E1_2", "categorical", "living", "Можуть оплатити тижневу відпустку", "Can afford a one-week holiday", "Усі домогосподарства", "All households"),
    "payment_arrears": Variable("E1_3", "categorical", "living", "Можуть розрахуватися з простроченими платежами", "Can settle payment arrears", "Усі домогосподарства", "All households"),
    "protein_meal": Variable("E1_4", "categorical", "living", "Можуть дозволити поживну їжу через день", "Can afford a protein-rich meal every other day", "Усі домогосподарства", "All households"),
    "warm_home": Variable("E1_5", "categorical", "living", "Можуть підтримувати житло теплим", "Can keep the home adequately warm", "Усі домогосподарства", "All households"),
}


COMMON_CATEGORY_LABELS: dict[str, tuple[str, str]] = {
    "yes": ("Так", "Yes"),
    "no": ("Ні", "No"),
    "unknown": ("Не знаю", "Don't know"),
    "жіноча": ("Жінки", "Women"),
    "чоловіча": ("Чоловіки", "Men"),
    "urban": ("Міська місцевість", "Urban area"),
    "rural": ("Сільська місцевість", "Rural area"),
    "Так": ("Так", "Yes"),
    "Ні": ("Ні", "No"),
    "Не знаю": ("Не знаю", "Don't know"),
}


REGION_EN = {
    "Вінницька обл.": "Vinnytsia",
    "Волинська обл.": "Volyn",
    "Дніпропетровська обл.": "Dnipropetrovsk",
    "Донецька обл.": "Donetsk",
    "Житомирська обл.": "Zhytomyr",
    "Закарпатська обл.": "Zakarpattia",
    "Запорізька обл.": "Zaporizhzhia",
    "Івано-Франківська обл.": "Ivano-Frankivsk",
    "Київська (без м.Києва) обл.": "Kyiv region",
    "Кіровоградська обл.": "Kirovohrad",
    "Львівська обл.": "Lviv",
    "м.Київ": "Kyiv city",
    "Миколаївська обл.": "Mykolaiv",
    "Одеська обл.": "Odesa",
    "Полтавська обл.": "Poltava",
    "Рівненська обл.": "Rivne",
    "Сумська обл.": "Sumy",
    "Тернопільська обл.": "Ternopil",
    "Харківська обл.": "Kharkiv",
    "Херсонська обл.": "Kherson",
    "Хмельницька обл.": "Khmelnytskyi",
    "Черкаська обл.": "Cherkasy",
    "Чернівецька обл.": "Chernivtsi",
    "Чернігівська обл.": "Chernihiv",
}


CATEGORY_EN: dict[str, str] = {
    **REGION_EN,
    "Перебуваю в зареєстрованому шлюбі": "Married (registered)",
    "Перебуваю в незареєстрованому шлюбі": "Living with a partner",
    "Ніколи не перебував(ла) у шлюбі": "Never married",
    "Розлучений(а) (зареєстроване розірвання шлюбу)": "Divorced",
    "Розійшовся(лася) (незареєстроване розірвання шлюбу)": "Separated",
    "Удівець, удова": "Widowed",
    "Працював (в тому числі займалась/вся бізнесом, будь-чим, що приносить дохід)": "Worked for income",
    "Не працював": "Did not work",
    "Не працював тимчасово (не більше місяця), хоча мав роботу або бізнес": "Temporarily absent from work",
    "Працював вдома (на власній земельній ділянці) для задоволення потреб членів родини без оплати": "Unpaid work for household needs",
    "Заробітна плата (чистий дохід, без податків і аліментів)": "Net salary",
    "Пенсія": "Pension",
    "Не маю індивідуального доходу": "No personal income",
    "Державна соціальна допомога": "State social assistance",
    "Підприємницький доход": "Business income",
    "Стипендія": "Scholarship",
    "Допомога по безробіттю": "Unemployment benefit",
    "Інше": "Other",
    "Другий (магістерський) рівень вищої освіти": "Master's degree",
    "Базова середня (базова загальна середня) освіта": "Lower secondary education",
    "Професійна (професійно-технічна) освіта": "Vocational education",
    "Дошкільна освіта": "Pre-primary education",
    "Повна загальна середня (профільна середня) освіта": "Upper secondary education",
    "Перший (бакалаврський) рівень вищої освіти (базова вища)": "Bachelor's degree",
    "Початкова (початкова загальна) освіта": "Primary education",
    "Початковий рівень (короткий цикл) вищої освіти (неповна вища)": "Short-cycle tertiary education",
    "Не має освіти, але вміє читати та писати": "No formal education; literate",
    "Не має освіти та не вміє читати та писати": "No formal education; not literate",
    "Науковий ступінь": "Research degree",
    "Проживав ще до початку гібридної війни з рф на Сході України (до 20 лютого 2014 року)": "Lived here before 20 February 2014",
    "Переїхав після початку гібридної війни з рф на Сході України (після 20 лютого 2014 року)": "Moved here after 20 February 2014",
    "Не знаю / не пам’ятаю": "Don't know / don't remember",
    "Робота, за яку платять заробітну плату, оформлена трудовим договором/ трудовою книжкою, контрактом чи іншим документом": "Formal salaried work",
    "Робота за зарплату без оформлення документів, а на основі усної домовленості": "Informal salaried work",
    "Працював індивідуально на себе заради отримання доходу без оформлення підприємства та без інших осіб (ремонти, консульта": "Unregistered self-employment",
    "Власник або керівник бізнесу, приватний підприємець (оформлене підприємство, ФОП)": "Registered business / sole proprietor",
    "Працював на сімейній земельній ділянці чи з худобою для вирощування на продаж": "Family farming for sale",
    "Охорона здоров'я та надання соціальної допомоги": "Health and social work",
    "Транспорт, складське господарство, поштова та кур'єрська діяльність": "Transport, warehousing and postal services",
    "Інші види економічної діяльності": "Other economic activities",
    "Освіта": "Education",
    "Сільське, лісове та рибне господарство": "Agriculture, forestry and fishing",
    "Оптова та роздрібна торгівля; ремонт автотранспортних засобів і мотоциклів": "Wholesale, retail and vehicle repair",
    "Промисловість": "Industry",
    "Діяльність у сфері адміністративного та допоміжного обслуговування": "Administrative and support services",
    "Фінансова та страхова діяльність": "Finance and insurance",
    "Інформація та телекомунікації": "Information and communications",
    "Будівництво": "Construction",
    "Державне управління й оборона; обов'язкове соціальне страхування": "Public administration, defence and social security",
    "Мистецтво, спорт, розваги та відпочинок": "Arts, sports and recreation",
    "Тимчасове розміщування й організація харчування": "Accommodation and food services",
    "Професійна, наукова та технічна діяльність": "Professional, scientific and technical activities",
    "Операції з нерухомим майном": "Real estate",
    "Інші професії": "Other occupations",
    "Робітники найпростіших професій": "Elementary occupations",
    "Працівники сфери торгівлі та послуг (продавці, касири та ін.)": "Service and sales workers",
    "Професіонали (наукові співробітники, математики, фізики, експерти, лікарі, гідрологи, маркетологи, інженери, тощо)": "Professionals",
    "Службовці, що виконують допоміжні функції (технічні службовці, клерки, секретарі)": "Clerical support workers",
    "Фахівці (техніки-будівельники, лаборанти, креслярі, диспетчери, інспектори, асистенти інженерів, брокери, агенти, тощо)": "Technicians and associate professionals",
    "Кваліфіковані робітники сільського та лісового господарств, риборозведення та рибальства": "Skilled agricultural, forestry and fishery workers",
    "Керівники (менеджери, управлінці) підприємств, установ, організацій та їх підрозділів": "Managers",
    "Робітники з обслуговування, експлуатації та контролювання за роботою технологічного устаткування, складання устаткування": "Plant and machine operators and assemblers",
    "Кваліфіковані робітники з інструментом": "Craft and related trades workers",
    "Значно зменшилися": "Decreased substantially",
    "Залишилися приблизно на одному рівні": "Stayed about the same",
    "Дещо зменшилися": "Decreased somewhat",
    "Дещо зросли": "Increased somewhat",
    "Значно зросли": "Increased substantially",
    "Не знає відповіді/ важко відповісти": "Don't know / difficult to answer",
    "Постійно відмовляли собі в найнеобхіднішому, крім харчування": "Regularly went without essentials other than food",
    "Не вдалося забезпечити навіть достатнє харчування": "Could not afford adequate food",
    "Було достатньо, але заощаджень не робили": "Enough for needs but no savings",
    "Було достатньо і робили заощадження": "Enough for needs and able to save",
    "Важко відповісти": "Difficult to answer",
    "Середній": "Average",
    "Нижче середнього": "Below average",
    "Вище середнього": "Above average",
    "Заощаджень не маємо": "No savings",
    "Маємо тривалі періоди фінансової стабільності": "Long periods of financial stability",
    "Не маємо фінансової стабільності": "No financial stability",
    "Маємо стабільні регулярні доходи": "Stable regular income",
    "Є власністю Вашої родини": "Owned by the family",
    "Ваша родина орендує": "Rented by the family",
    "Соціальне житло, надане для переселенців": "Social housing for displaced people",
    "Окрема квартира у багатоквартирному будинку": "Separate apartment in a multi-unit building",
    "Індивідуальний будинок (садиба)": "Detached house",
    "Частина індивідуального будинку": "Part of a detached house",
    "Комунальна квартира": "Communal apartment",
    "Гуртожиток": "Dormitory",
    "Модульне містечко": "Modular settlement",
    "Ні, потрібно більше площі": "No, more space is needed",
    "Так, площі достатньо": "Yes, space is sufficient",
    "Площі більше, ніж потрібно": "More space than needed",
}


def category_labels(value: str) -> dict[str, str]:
    if value in COMMON_CATEGORY_LABELS:
        uk, en = COMMON_CATEGORY_LABELS[value]
        return {"uk": uk, "en": en}
    return {"uk": value, "en": CATEGORY_EN.get(value, value)}


def find_one(pattern: str) -> Path:
    matches = sorted(RAW_DIR.glob(pattern))
    if len(matches) != 1:
        raise RuntimeError(f"Expected exactly one file matching {pattern!r}, found {len(matches)}")
    return matches[0]


def iter_records(path: Path, header_row: int) -> Iterable[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows)
    header = [str(v) if v is not None else "" for v in next(rows)]
    for values in rows:
        if not values or values[0] is None:
            continue
        yield {header[i]: values[i] if i < len(values) else None for i in range(len(header)) if header[i]}


def weighted_quantile(values: list[tuple[float, float]], q: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values, key=lambda item: item[0])
    target = q * sum(weight for _, weight in ordered)
    cumulative = 0.0
    for value, weight in ordered:
        cumulative += weight
        if cumulative >= target:
            return value
    return ordered[-1][0]


def weighted_share(records: list[dict[str, Any]], predicate: Callable[[dict[str, Any]], bool], eligible: Callable[[dict[str, Any]], bool] = lambda _: True) -> dict[str, Any]:
    base = [row for row in records if eligible(row)]
    matched = [row for row in base if predicate(row)]
    denominator = sum(row["weight"] for row in base)
    numerator = sum(row["weight"] for row in matched)
    return {
        "share": numerator / denominator if denominator else None,
        "weighted": numerator,
        "n": len(matched),
        "base_n": len(base),
    }


def age_band(age: float | None) -> str | None:
    if age is None:
        return None
    if age <= 2:
        return "0–2"
    if age <= 6:
        return "3–6"
    if age <= 13:
        return "7–13"
    if age <= 15:
        return "14–15"
    if age <= 17:
        return "16–17"
    if age <= 29:
        return "18–29"
    if age <= 59:
        return "30–59"
    return "60+"


def extract(path: Path, header_row: int, variables: dict[str, Variable], cluster_ids: dict[str, int] | None = None) -> tuple[list[dict[str, Any]], dict[str, int]]:
    output: list[dict[str, Any]] = []
    is_people = cluster_ids is not None
    person_numbers: dict[int, int] = defaultdict(int)
    ids = cluster_ids if cluster_ids is not None else {}
    required_sources = {definition.source for definition in variables.values()} | {"KEY_QUEST", "WGT_Fin_NEW", "IND_A15", "IND_A19", "IND_V1_1", "hh_size"}
    for source_row in iter_records(path, header_row):
        key = str(source_row["KEY_QUEST"])
        if key not in ids:
            ids[key] = len(ids) + 1
        row = {
            "cluster": ids[key],
            "weight": number(source_row.get("WGT_Fin_NEW")),
        }
        if is_people:
            person_numbers[row["cluster"]] += 1
            row["person"] = person_numbers[row["cluster"]]
        if row["weight"] is None:
            raise RuntimeError(f"Missing weight for household {key}")
        compact_source = {key: source_row.get(key) for key in required_sources}
        for name, definition in variables.items():
            value = definition.transform(compact_source) if definition.transform else compact_source.get(definition.source)
            row[name] = value
        output.append(row)
    return output, ids


def encode_dataset(records: list[dict[str, Any]], variables: dict[str, Variable], unit: str) -> dict[str, Any]:
    columns: dict[str, list[Any]] = {"cluster": [], "weight": []}
    if unit == "people":
        columns["person"] = []
    dictionaries: dict[str, list[dict[str, Any]]] = {}
    for name in variables:
        columns[name] = []

    for name, definition in variables.items():
        if definition.kind != "categorical":
            continue
        seen: dict[str, int] = {}
        entries: list[dict[str, Any]] = []
        for row in records:
            value = row[name]
            if value is None:
                continue
            text = str(value)
            if text not in seen:
                seen[text] = len(entries)
                entries.append({"value": text, "labels": category_labels(text)})
        dictionaries[name] = entries

    dictionary_indices = {
        name: {entry["value"]: index for index, entry in enumerate(entries)}
        for name, entries in dictionaries.items()
    }
    for row in records:
        columns["cluster"].append(row["cluster"])
        columns["weight"].append(round(row["weight"], 6))
        if unit == "people":
            columns["person"].append(row["person"])
        for name, definition in variables.items():
            value = row[name]
            if definition.kind == "categorical":
                columns[name].append(None if value is None else dictionary_indices[name][str(value)])
            else:
                columns[name].append(None if value is None else round(float(value), 6))

    return {
        "schema_version": 1,
        "unit": unit,
        "n": len(records),
        "columns": columns,
        "dictionaries": dictionaries,
    }


def variable_metadata(variables: dict[str, Variable]) -> list[dict[str, Any]]:
    return [
        {
            "id": name,
            "type": definition.kind,
            "topic": definition.topic,
            "labels": {"uk": definition.label_uk, "en": definition.label_en},
            "universe": {"uk": definition.universe_uk, "en": definition.universe_en},
        }
        for name, definition in variables.items()
    ]


def build_overview(people: list[dict[str, Any]], households: list[dict[str, Any]]) -> dict[str, Any]:
    age_sex = defaultdict(lambda: {"weighted": 0.0, "n": 0})
    for row in people:
        band = age_band(row["age"])
        if band is None or row["sex"] is None:
            continue
        key = (band, row["sex"])
        age_sex[key]["weighted"] += row["weight"]
        age_sex[key]["n"] += 1

    quantiles = [0.1, 0.25, 0.5, 0.75, 0.9]
    total_income = [(row["hh_income_total"], row["weight"]) for row in households if row["hh_income_total"] is not None]
    per_capita = [(row["hh_income_per_capita"], row["weight"]) for row in households if row["hh_income_per_capita"] is not None]

    work_status = defaultdict(lambda: {"weighted": 0.0, "n": 0})
    arrangement = defaultdict(lambda: {"weighted": 0.0, "n": 0})
    for row in people:
        if row["age"] is not None and row["age"] >= 15 and row["work_status"] is not None:
            work_status[row["work_status"]]["weighted"] += row["weight"]
            work_status[row["work_status"]]["n"] += 1
        if row["work_arrangement"] is not None:
            arrangement[row["work_arrangement"]]["weighted"] += row["weight"]
            arrangement[row["work_arrangement"]]["n"] += 1

    deprivation_variables = ["unexpected_expense", "annual_holiday", "protein_meal", "warm_home"]
    deprivation = []
    for variable in deprivation_variables:
        groups = {}
        for key, has_children in (("with_children", True), ("without_children", False)):
            subset = [row for row in households if (row["children_count"] > 0) == has_children and row[variable] is not None]
            no_value = "Ні"
            groups[key] = weighted_share(subset, lambda row, v=variable: row[v] == no_value)
        deprivation.append({"variable": variable, **groups})

    return {
        "schema_version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "totals": {
            "people": sum(row["weight"] for row in people),
            "people_n": len(people),
            "households": sum(row["weight"] for row in households),
            "households_n": len(households),
            "average_household_size": sum(row["weight"] for row in people) / sum(row["weight"] for row in households),
            "households_with_children_share": sum(row["weight"] for row in households if row["children_count"] and row["children_count"] > 0) / sum(row["weight"] for row in households),
        },
        "age_sex": [
            {"age_band": band, "sex": sex, **values}
            for (band, sex), values in age_sex.items()
        ],
        "income": {
            "quantiles": [
                {
                    "p": q,
                    "total": weighted_quantile(total_income, q),
                    "per_capita": weighted_quantile(per_capita, q),
                }
                for q in quantiles
            ],
            "n": len(total_income),
        },
        "employment": {
            "status": [{"category": category, **values} for category, values in work_status.items()],
            "arrangement": [{"category": category, **values} for category, values in arrangement.items()],
        },
        "deprivation": deprivation,
    }


def validate(people: list[dict[str, Any]], households: list[dict[str, Any]]) -> None:
    if len(people) != 18837:
        raise AssertionError(f"Expected 18,837 people, got {len(people):,}")
    if len(households) != 8023:
        raise AssertionError(f"Expected 8,023 households, got {len(households):,}")
    if len({row["cluster"] for row in households}) != 8023:
        raise AssertionError("Household cluster ids are not unique")
    if not {row["cluster"] for row in people}.issubset({row["cluster"] for row in households}):
        raise AssertionError("A person references a household missing from the household file")
    if len({(row["cluster"], row["person"]) for row in people}) != len(people):
        raise AssertionError("Person keys are not unique within household clusters")

    people_total = sum(row["weight"] for row in people)
    household_total = sum(row["weight"] for row in households)
    if abs(people_total - 31_143_332) > 2:
        raise AssertionError(f"Unexpected population weight total: {people_total:,.3f}")
    if abs(household_total - 13_574_519) > 2:
        raise AssertionError(f"Unexpected household weight total: {household_total:,.3f}")

    weights_by_cluster = {row["cluster"]: row["weight"] for row in households}
    if any(abs(row["weight"] - weights_by_cluster[row["cluster"]]) > 1e-6 for row in people):
        raise AssertionError("Person and household weights disagree within a cluster")

    forbidden = {"KEY_QUEST", "RAJON", "contact_phone", "IND_B6_1", "IND_B4_1_1_1"}
    public_fields = set().union(*(row.keys() for row in people[:1] + households[:1]))
    if public_fields & forbidden:
        raise AssertionError(f"Forbidden fields leaked: {public_fields & forbidden}")


def write_json(path: Path, payload: Any, pretty: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(
            payload,
            handle,
            ensure_ascii=False,
            indent=2 if pretty else None,
            separators=None if pretty else (",", ":"),
        )
        handle.write("\n")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true", help="Validate source data without writing outputs")
    args = parser.parse_args()

    people_path = find_one(PEOPLE_GLOB)
    household_path = find_one(HOUSEHOLD_GLOB)

    households, cluster_ids = extract(household_path, 5, HOUSEHOLD_VARIABLES)
    people, _ = extract(people_path, 4, PEOPLE_VARIABLES, cluster_ids)
    validate(people, households)
    overview = build_overview(people, households)
    encoded_people = encode_dataset(people, PEOPLE_VARIABLES, "people")
    encoded_households = encode_dataset(households, HOUSEHOLD_VARIABLES, "households")

    metadata = {
        "schema_version": 1,
        "title": {"uk": "Україна у даних — SESH 2023/24", "en": "Ukraine in Data — SESH 2023/24"},
        "fieldwork": FIELDWORK,
        "territory": {
            "uk": "Підконтрольна територія України; без АР Крим і Луганської області, у Донецькій, Запорізькій та Херсонській областях — лише підконтрольні території.",
            "en": "Government-controlled territory of Ukraine; excluding Crimea and Luhansk region, with only government-controlled areas in Donetsk, Zaporizhzhia and Kherson regions.",
        },
        "source": {
            "name": "Household Socio-Economic Status Survey (HSESS/SESH)",
            "url": SOURCE_URL,
            "license": LICENSE,
            "credits": "Institute of Demography and Quality of Life Problems of the NAS of Ukraine; Ukrainian Center for Social Reforms; UNICEF; Ministry of Social Policy of Ukraine; BMZ/KfW",
        },
        "units": {
            "people": {"n": len(people), "weight_total": sum(row["weight"] for row in people), "variables": variable_metadata(PEOPLE_VARIABLES), "dictionaries": encoded_people["dictionaries"]},
            "households": {"n": len(households), "weight_total": sum(row["weight"] for row in households), "variables": variable_metadata(HOUSEHOLD_VARIABLES), "dictionaries": encoded_households["dictionaries"]},
        },
        "reliability": {"suppress_below": 10, "warn_below": 30},
    }

    if args.check:
        print(json.dumps({"people": len(people), "households": len(households), "people_weight": overview["totals"]["people"], "household_weight": overview["totals"]["households"]}, ensure_ascii=False, indent=2))
        return

    write_json(OUT_DIR / "people.json", encoded_people)
    write_json(OUT_DIR / "households.json", encoded_households)
    write_json(OUT_DIR / "overview.json", overview, pretty=True)
    write_json(OUT_DIR / "metadata.json", metadata, pretty=True)
    print(f"Built SESH dashboard data: {len(people):,} people, {len(households):,} households")


if __name__ == "__main__":
    main()
